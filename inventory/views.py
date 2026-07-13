import json
from multiprocessing import context
from urllib import request
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from products.models import Product, Category
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from .models import Purchase
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter



def product_list(request):
    query = request.GET.get('q', '')
    category_id = request.GET.get('category', '')
    stock_status = request.GET.get('stock', '')
    
    products = Product.objects.all().select_related('category')
    categories = Category.objects.all()
    
    if query:
        from django.db.models import Q
        products = products.filter(
            Q(name__icontains=query) |
            Q(sku__icontains=query) |
            Q(color__icontains=query) |
            Q(size__icontains=query)
        )
    
    if category_id:
        products = products.filter(category_id=category_id)
        
    if stock_status:
        if stock_status == 'low-stock':
            products = products.filter(quantity__lte=5)
        elif stock_status == 'in-stock':
            products = products.filter(quantity__gt=5)
            
    context = {
        'products': products,
        'categories': categories,
        'query': query,
        'selected_category': int(category_id) if category_id.isdigit() else '',
        'selected_stock': stock_status,
    }
    return render(request, 'products.html', context)

def low_stock_items(request):
    """Display only low stock items"""
    low_stock_products = Product.objects.filter(quantity__lte=5).select_related('category').order_by('quantity')
    
    context = {
        'products': low_stock_products,
        'total_low_stock': low_stock_products.count(),
    }
    return render(request, 'inventory/low_stock.html', context)



@login_required
def export_purchase_excel(request):

    if not request.user.is_superuser:
        raise PermissionDenied("Only administrators can export purchase reports.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Report"

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "HAMRO MART - PURCHASE REPORT"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = [
        "Purchase ID",
        "Product",
        "Supplier",
        "Purchase Date",
        "Purchase Price",
        "Cost Price",
    ]

    fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill

    purchases = Purchase.objects.select_related("product").order_by("-purchase_date")

    row = 4
    total_purchase = 0

    for purchase in purchases:

        ws.cell(row=row, column=1).value = purchase.id
        ws.cell(row=row, column=2).value = purchase.product.name
        ws.cell(row=row, column=3).value = purchase.supplier_name
        ws.cell(row=row, column=4).value = purchase.purchase_date.strftime("%d-%m-%Y")
        ws.cell(row=row, column=5).value = float(purchase.purchase_price)
        ws.cell(row=row, column=6).value = float(purchase.product.cost_price)

        total_purchase += float(purchase.purchase_price)

        row += 1

    ws.cell(row=row + 1, column=4).value = "Total Purchase"
    ws.cell(row=row + 1, column=4).font = Font(bold=True)

    ws.cell(row=row + 1, column=5).value = total_purchase
    ws.cell(row=row + 1, column=5).font = Font(bold=True)

    # Auto Width
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        length = 0

        for cell in ws[letter]:
            if cell.coordinate in ws.merged_cells:
                continue

            if cell.value:
                length = max(length, len(str(cell.value)))

        ws.column_dimensions[letter].width = length + 5

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="HamroMart_Purchase_Report.xlsx"'

    wb.save(response)

    return response