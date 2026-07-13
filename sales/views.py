import json
import re
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from products.models import Product
from customers.models import Customer
from .models import Sale, SaleItem,DiscountScheme
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

PHONE_NUMBER_RE = re.compile(r'^(97|98)\d{8}$')

STAFF_POOL_MAX = 25   # max units visible to staff at once
STAFF_LOW_THRESHOLD = 5  # trigger auto-restore when staff pool hits this

def is_valid_phone_number(phone):
    phone = (phone or '').strip()
    return bool(PHONE_NUMBER_RE.fullmatch(phone))

@login_required
def sales_report(request):
    sales = Sale.objects.all().order_by('-timestamp')

    context = {
        "sales": sales,
    }

    return render(request, "sales/sales_report.html", context)

@login_required
def pos_view(request):
    # Both admin and staff can access POS
    if not ( request.user.groups.filter(name='Cashier').exists()):
        raise PermissionDenied("You do not have permission to access the POS.")
    # Staff only sees products with remaining staff_quantity
    is_admin = request.user.is_superuser
    if is_admin:
        products = Product.objects.filter(quantity__gt=0).order_by('name')
    else:
        products = Product.objects.filter(staff_quantity__gt=0).order_by('name')
    scheme=DiscountScheme.objects.filter(is_active=True).first()
    return render(request, 'pos/pos.html', {
        'products': products,
        'is_admin': is_admin,
        'staff_low_threshold': STAFF_LOW_THRESHOLD,
        'scheme': scheme,
    })

    
@login_required
def get_customer_history(request):
    phone = (request.GET.get('phone') or '').strip()
    try:
        customer = Customer.objects.get(phone=phone)
        sales = Sale.objects.filter(customer=customer).order_by('-timestamp')[:10]
        history = [{
            "id": s.id,
            "date": s.timestamp.strftime("%Y-%m-%d"),
            "total": float(s.total),
            "bill_no": s.bill_no,
        } for s in sales]
        return JsonResponse({"status": "success", "name": customer.name, "history": history})
    except Customer.DoesNotExist:
        return JsonResponse({"status": "not_found"})
    
@login_required
def customer_history_view(request):
    return render(request, 'pos/customer_history.html')


@login_required
def process_sale(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            phone = (data.get('phone') or '').strip()
            if not is_valid_phone_number(phone):
                return JsonResponse(
                    {"status": "error", "message": "Phone number must be exactly 10 digits and start with 97 or 98."},
                    status=400,
                )

            customer = None
            if phone:
                customer, _ = Customer.objects.get_or_create(
                    phone=phone,
                    defaults={'name': data.get('name', 'Walk-in Customer')}
                )

            sale = Sale.objects.create(
                customer=customer,
                subtotal=data['subtotal'],
                discount=data.get('discount', 0),
                total=data['total'],
                payment_method='Cash',
                payment_status='Completed',
                cashier=request.user,
            )

            is_admin = request.user.is_superuser

            updated_products = []

            for item in data['cart']:
                product = Product.objects.get(id=item['id'])
                qty_sold = int(item['qty'])

                # Check available stock
                if request.user.is_superuser:
                    available = product.quantity
                else:
                    available = product.staff_quantity

                if qty_sold > available:
                    sale.delete()
                    return JsonResponse(
                        {
                            "status": "error",
                            "message": f"Only {available} unit(s) of {product.name} are available."
                        },
                        status=400,
                    )

                SaleItem.objects.create(
                    sale=sale,
                    product=product,
                    quantity=qty_sold,
                    price=item['price'],
                )

                # Deduct from master stock (always)
                product.quantity = max(0, product.quantity - qty_sold)

                if is_admin:
                    # Admin sells from master — also bring staff_quantity in sync
                    product.staff_quantity = min(STAFF_POOL_MAX, product.quantity)
                else:
                    # Staff sells from their pool
                    product.staff_quantity = max(0, product.staff_quantity - qty_sold)

                product.save()

                updated_products.append({
                    "id": product.id,
                    "quantity": product.quantity,
                    "staff_quantity": product.staff_quantity,
                })

                # Auto-restore staff pool if running low and master stock available
                if (
                    product.staff_quantity <= STAFF_LOW_THRESHOLD
                    and product.quantity > product.staff_quantity
                ):
                    restore_to = min(STAFF_POOL_MAX, product.quantity)
                    product.staff_quantity = restore_to
                    product.save()

            return JsonResponse({
                "status": "success",
                "payment_method": "Cash",
                "payment_status": "Completed",
                "sale_id": sale.id,
                "bill_no": sale.bill_no,
                "cashier": sale.cashier.username if sale.cashier else None,
            })
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)


@login_required
def receipt_detail(request, pk):
    if not (request.user.is_superuser or request.user.groups.filter(name='Cashier').exists()):
        raise PermissionDenied("Only administrators and cashiers are allowed to view receipt details.")
    sale = get_object_or_404(Sale, pk=pk)
    return render(request, 'receipt/receipt_detail.html', {'sale': sale})
    # return JsonResponse ({"success": true,"bill_no": sale.bill_no})


@login_required
def export_sales_excel(request):
    # Only Admin can export sales report
    if not request.user.is_superuser:
        raise PermissionDenied("Only administrators can export sales reports.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Report Title
    ws.merge_cells("A1:I1")
    ws["A1"] = "Hamro Mart Sales Report"
    ws["A1"].font = Font(size=16, bold=True)

    # Header Row
    headers = [
        "Bill No",
        "Customer",
        "Cashier",
        "Date & Time",
        "Subtotal",
        "Discount",
        "Total",
        "Payment Method",
        "Payment Status",
    ]

    fill = PatternFill(start_color="4F81BD",
                       end_color="4F81BD",
                       fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = fill

    sales = Sale.objects.select_related(
        "customer",
        "cashier"
    ).order_by("-timestamp")

    row = 4
    total_sales = 0

    for sale in sales:
        ws.cell(row=row, column=1).value = sale.bill_no
        ws.cell(row=row, column=2).value = sale.customer.name if sale.customer else "Walk-in Customer"
        ws.cell(row=row, column=3).value = sale.cashier.username if sale.cashier else "-"
        ws.cell(row=row, column=4).value = sale.timestamp.strftime("%d-%m-%Y %H:%M")
        ws.cell(row=row, column=5).value = float(sale.subtotal)
        ws.cell(row=row, column=6).value = float(sale.discount)
        ws.cell(row=row, column=7).value = float(sale.total)
        ws.cell(row=row, column=8).value = sale.payment_method
        ws.cell(row=row, column=9).value = sale.payment_status

        total_sales += float(sale.total)
        row += 1

    # Summary
    row += 2
    ws.cell(row=row, column=6).value = "Grand Total"
    ws.cell(row=row, column=6).font = Font(bold=True)
    ws.cell(row=row, column=7).value = total_sales
    ws.cell(row=row, column=7).font = Font(bold=True)

    # Auto-size columns
    from openpyxl.utils import get_column_letter

    for col in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col)

        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
        try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        except Exception:
            pass

    ws.column_dimensions[column_letter].width = max_length + 5

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="HamroMart_Sales_Report.xlsx"'

    wb.save(response)
    return response